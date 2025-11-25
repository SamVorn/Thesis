import os
import json
import csv
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import yaml
from pathlib import Path

# ---------------------------
# LOAD SURVEY FILE (ANY FORMAT)
# ---------------------------
def load_survey(file_path: str) -> dict:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            raw_data = json.load(f)
        return _normalize_json(raw_data, file_path)

    elif ext == ".csv":
        return _parse_csv(file_path)

    elif ext in (".yml", ".yaml"):
        return _parse_yaml(file_path)

    elif ext == ".xml":
        return _parse_xml(file_path)

    elif ext == ".xlsx":
        return _parse_xlsx(file_path)

    elif ext == ".txt":
        return _parse_txt(file_path)

    else:
        raise ValueError(f"Unsupported survey file type: {ext}")


# ---------------------------
# LOAD RULES (JSON/YAML/TXT)
# ---------------------------
def load_rules(file_path: str) -> dict:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    with open(file_path, "r", encoding="utf-8") as f:
        if ext in (".json", ".yml", ".yaml"):
            return yaml.safe_load(f) if ext in (".yml", ".yaml") else json.load(f)
        elif ext == ".txt":
            rules = {}
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if ":" in line:
                    label, pattern = line.split(":", 1)
                    rules[label.strip()] = pattern.strip()
            return rules
        else:
            raise ValueError(f"Unsupported rule file type: {ext}")


# ---------------------------
# JSON PARSING
# ---------------------------
def _normalize_json(data, file_path: str) -> dict:
    keys = ["responses", "questions", "items", "entries"]
    responses = []
    response_list = next((data.get(k) for k in keys if k in data), [])
    if not response_list and "sections" in data:
        for section in data["sections"]:
            response_list.extend(section.get("questions", []))

    for i, r in enumerate(response_list, start=1):
        responses.append({
            "question_id": r.get("question_id") or r.get("id") or i,
            "question_text": r.get("question_text") or r.get("text") or r.get("question") or "",
            "answer": r.get("answer") or r.get("response") or r.get("value") or ""
        })
    survey_id = data.get("survey_id") or data.get("id") or os.path.basename(file_path)
    return {"survey_id": survey_id, "responses": responses}


# ---------------------------
# CSV PARSING
# ---------------------------
def _parse_csv(file_path: str) -> dict:
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        responses = []
        for i, row in enumerate(reader, start=1):
            responses.append({
                "question_id": row.get("question_id") or row.get("id") or i,
                "question_text": row.get("question_text") or row.get("question") or "",
                "answer": row.get("answer") or row.get("response") or ""
            })
    return {"survey_id": os.path.basename(file_path), "responses": responses}


# ---------------------------
# YAML PARSING
# ---------------------------
def _parse_yaml(file_path: str) -> dict:
    with open(file_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return _normalize_json(data, file_path)


# ---------------------------
# TXT PARSING
# ---------------------------
def _parse_txt(file_path: str) -> dict:
    responses = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    i = 0
    counter = 1
    while i < len(lines):
        line = lines[i]
        if ":" in line:
            parts = line.split(":", 1)
            responses.append({
                "question_id": counter,
                "question_text": parts[0].strip(),
                "answer": parts[1].strip()
            })
            i += 1
        else:
            question = lines[i]
            answer = lines[i+1] if i+1 < len(lines) else ""
            responses.append({
                "question_id": counter,
                "question_text": question,
                "answer": answer
            })
            i += 2
        counter += 1
    return {"survey_id": os.path.basename(file_path), "responses": responses}


# ---------------------------
# XML PARSING
# ---------------------------
def _parse_xml(file_path: str) -> dict:
    tree = ET.parse(file_path)
    root = tree.getroot()

    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    survey_id = root.attrib.get("id", os.path.basename(file_path))
    responses = []

    # --- FIX: ElementTree does NOT support OR in predicates ---
    # So we gather all matching containers manually
    tags_to_search = [f"{ns}answer", f"{ns}response", f"{ns}text"]

    response_containers = []
    for tag in tags_to_search:
        for el in root.findall(f".//{tag}/.."):  # parent of the element
            if el not in response_containers:
                response_containers.append(el)

    # Fallback: if nothing matched, just treat direct children as responses
    if not response_containers:
        response_containers = list(root)

    for i, container in enumerate(response_containers, start=1):
        q_id = (
            container.attrib.get("id")
            or container.attrib.get("question_id")
            or i
        )

        q_text = (
            container.findtext(f"{ns}question_text")
            or container.findtext(f"{ns}text")
            or container.tag.replace(ns, "").replace("_", " ").title()
        )

        answer = (
            container.findtext(f"{ns}answer")
            or container.findtext(f"{ns}response")
            or container.findtext(f"{ns}value")
            or (container.text or "").strip()
        )

        if not answer and container.attrib:
            answer = next(iter(container.attrib.values()))

        if q_text or answer:
            responses.append({
                "question_id": q_id,
                "question_text": q_text,
                "answer": answer
            })

    return {"survey_id": survey_id, "responses": responses}


# ---------------------------
# XLSX PARSING
# ---------------------------
def _parse_xlsx(file_path: str) -> dict:
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return {"survey_id": os.path.basename(file_path), "responses": []}

        # Convert all cells to string
        data_rows = [[str(c).strip() if c is not None else "" for c in row] for row in all_rows]

        # Detect header row
        header = data_rows[0]
        if sum(1 for c in header if c) >= len(header)/2:
            headers = [c.lower() for c in header]
            data_start = 1
        else:
            headers = [f"col_{i}" for i in range(len(header))]
            data_start = 0

        # Guess question/answer columns
        q_col = a_col = id_col = None
        for i, h in enumerate(headers):
            if "question" in h or "prompt" in h or "text" in h:
                q_col = i
            if "answer" in h or "response" in h or "value" in h:
                a_col = i
            if "id" in h or "qid" in h:
                id_col = i
        if q_col is None:
            q_col = 0
        if a_col is None:
            a_col = 1 if len(headers) > 1 else 0

        # Extract responses
        responses = []
        counter = 1
        for row in data_rows[data_start:]:
            if all(c == "" for c in row):
                continue
            question_id = row[id_col] if id_col is not None and id_col < len(row) and row[id_col] else str(counter)
            question_text = row[q_col] if q_col < len(row) else ""
            answer = row[a_col] if a_col < len(row) else ""
            responses.append({
                "question_id": str(question_id),
                "question_text": str(question_text),
                "answer": str(answer)
            })
            counter += 1

        return {"survey_id": os.path.basename(file_path), "responses": responses}

    except Exception as e:
        raise Exception(f"Error parsing XLSX file: {e}")
